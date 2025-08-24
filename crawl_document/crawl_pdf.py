from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
from openpyxl import load_workbook
import time
import sys

def download_pdfs_from_excel(excel_file, output_folder, download_column_name='download-href'):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    prefs = {
        "download.default_directory": output_folder,
        "plugins.always_open_pdf_externally": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)

    wb = load_workbook(excel_file)
    ws = wb.active

    col_index = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and cell.value.strip().lower() == download_column_name.lower():
            col_index = idx
            break
    if col_index is None:
        driver.quit()
        raise ValueError(f"Không tìm thấy cột '{download_column_name}'")

    for row in ws.iter_rows(min_row=2):
        url = row[col_index-1].value
        if url:
            url = url.strip()
            driver.get(url)
            time.sleep(5)
            print(f"Đã tải xuống: {url}")

    driver.quit()
    print(f"Tất cả PDF đã được lưu vào thư mục: {output_folder}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Vui lòng chạy lệnh: python crawl_pdf.py <tên_file_excel>")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_folder = os.path.join(os.getcwd(), "Doccument_mdpi_Medical_Journal")
    download_pdfs_from_excel(excel_file, output_folder)
